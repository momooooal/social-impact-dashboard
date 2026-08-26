from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

APP_NAME = "SocialImpactCollector"
PLATFORMS = ("facebook", "instagram", "threads")

TOPICS = {
    "registration": ("報名方式", ["怎麼報名", "如何報名", "報名網址", "報名連結", "哪裡報名", "登記", "報名方式", "register", "registration"]),
    "capacity": ("名額／候補", ["額滿", "名額", "候補", "備取", "還有名額", "順位", "waitlist", "full"]),
    "eligibility": ("資格／參加對象", ["資格", "可以參加", "能參加", "年齡", "幾歲", "對象", "身分", "戶籍", "限制", "親子", "兒童", "長者", "eligible"]),
    "time": ("時間／日期", ["幾點", "時間", "日期", "哪一天", "哪天", "開始時間", "報到", "幾號", "when", "time", "date"]),
    "location": ("地點／交通／停車", ["地點", "在哪", "地址", "交通", "停車", "捷運", "公車", "怎麼去", "會場", "parking", "location", "address"]),
    "fee": ("費用／付款／退費", ["費用", "多少錢", "免費", "繳費", "付款", "退費", "退款", "price", "fee", "refund"]),
    "program": ("活動內容／流程／規則", ["活動內容", "流程", "怎麼玩", "賽程", "規則", "辦法", "行程", "比賽方式", "組別", "rule", "schedule"]),
    "equipment": ("裝備／服裝／材料", ["帶什麼", "裝備", "服裝", "穿什麼", "材料", "自備", "攜帶", "equipment", "wear"]),
    "weather": ("天候／延期／取消", ["下雨", "颱風", "天氣", "延期", "取消", "停辦", "雨天", "照常", "rain", "weather", "cancel"]),
    "award": ("獎項／成績／證明", ["獎金", "獎品", "成績", "名次", "證書", "完賽", "獎牌", "prize", "result"]),
    "other": ("其他", []),
}

METRIC_LABELS = {
    "followers_end": ["追蹤者", "粉絲人數", "followers", "followers count", "follower count"],
    "views": ["瀏覽次數", "觀看次數", "觀看", "views", "view count"],
    "reach": ["觸及人數", "觸及", "reach", "accounts reached"],
    "content_interactions": ["內容互動", "互動次數", "互動", "content interactions", "interactions"],
    "profile_views": ["造訪次數", "個人檔案瀏覽", "粉絲專頁瀏覽", "profile visits", "page visits", "visits"],
    "inbox_conversations": ["開始的對話", "訊息對話", "新對話", "messaging conversations started", "conversations started"],
    "inbox_messages": ["收到的訊息", "訊息則數", "messages received", "messages"],
}

HEADER_ALIASES = {
    "published_at": ["發布時間", "發佈時間", "建立時間", "post date", "published", "date", "日期", "created time"],
    "text": ["文案", "貼文文字", "標題", "description", "caption", "post message", "message", "內容"],
    "url": ["網址", "永久連結", "permalink", "url", "link"],
    "content_type": ["內容類型", "貼文類型", "content type", "post type", "type", "媒體類型"],
    "views": ["觀看", "觀看次數", "瀏覽次數", "views", "video views", "plays", "impressions"],
    "reach": ["觸及", "觸及人數", "reach", "accounts reached"],
    "likes": ["讚", "心情", "反應", "likes", "reactions"],
    "comments": ["留言", "回覆", "comments", "replies"],
    "shares": ["分享", "轉發", "shares", "reposts"],
    "saves": ["收藏", "saves", "saved"],
    "clicks": ["連結點擊", "點擊", "link clicks", "clicks"],
}


def app_data_dir() -> Path:
    if os.name == "nt":
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
    else:
        base = Path.home() / ".local" / "share"
    p = base / APP_NAME
    p.mkdir(parents=True, exist_ok=True)
    return p


def profile_dir() -> Path:
    p = app_data_dir() / "browser-profile"
    p.mkdir(parents=True, exist_ok=True)
    return p


def runtime_dir() -> Path:
    p = app_data_dir() / "runtime"
    p.mkdir(parents=True, exist_ok=True)
    return p


def raw_dir() -> Path:
    p = app_data_dir() / "raw"
    p.mkdir(parents=True, exist_ok=True)
    return p


def private_dir() -> Path:
    p = app_data_dir() / "private"
    p.mkdir(parents=True, exist_ok=True)
    return p


def config_path() -> Path:
    return app_data_dir() / "config.json"


def project_dir() -> Path:
    return Path(__file__).resolve().parent.parent


def default_config_path() -> Path:
    return Path(__file__).resolve().parent / "config.default.json"


def ensure_config() -> Path:
    dst = config_path()
    if not dst.exists():
        shutil.copy2(default_config_path(), dst)
    return dst


def load_config() -> dict[str, Any]:
    ensure_config()
    return json.loads(config_path().read_text(encoding="utf-8"))


def save_config(cfg: dict[str, Any]) -> None:
    config_path().write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def output_data_path(cfg: dict[str, Any] | None = None) -> Path:
    cfg = cfg or load_config()
    repo = str(cfg.get("dashboard_repo_path", "")).strip()
    if repo:
        path = Path(repo).expanduser().resolve() / "data" / "collector-data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        return path
    return project_dir() / "data" / "collector-data.json"


def load_output_data(cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    p = output_data_path(cfg)
    if p.exists():
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            d = {}
    else:
        d = {}
    d.setdefault("meta", {"schema_version": 1, "source": "local-helper", "updated_at": ""})
    d.setdefault("monthly", [])
    d.setdefault("content", [])
    d.setdefault("inquiry_aggregates", [])
    d.setdefault("collector_status", [])
    return d


def save_output_data(data: dict[str, Any], cfg: dict[str, Any] | None = None) -> Path:
    data.setdefault("meta", {})["updated_at"] = datetime.now().astimezone().isoformat()
    p = output_data_path(cfg)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def normalize_text(s: Any) -> str:
    return re.sub(r"[\s\W_]+", "", str(s or "").lower(), flags=re.UNICODE)


def parse_human_number(value: str) -> int | float | None:
    s = str(value or "").strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    m = re.search(r"(-?\d+(?:\.\d+)?)\s*(萬|万|k|K|m|M)?", s)
    if not m:
        return None
    num = float(m.group(1))
    suffix = m.group(2)
    if suffix in ("萬", "万"):
        num *= 10000
    elif suffix in ("k", "K"):
        num *= 1000
    elif suffix in ("m", "M"):
        num *= 1_000_000
    return int(round(num)) if abs(num - round(num)) < 1e-9 else num


def extract_metrics_from_text(text: str) -> dict[str, int | float]:
    lines = [re.sub(r"\s+", " ", x).strip() for x in text.splitlines() if x.strip()]
    out: dict[str, int | float] = {}
    for key, labels in METRIC_LABELS.items():
        candidates: list[int | float] = []
        for i, line in enumerate(lines):
            low = line.lower()
            for label in labels:
                pos = low.find(label.lower())
                if pos < 0:
                    continue
                fragments = [line[pos + len(label):], line[:pos]]
                if i + 1 < len(lines): fragments.append(lines[i + 1])
                if i + 2 < len(lines): fragments.append(lines[i + 2])
                for frag in fragments:
                    v = parse_human_number(frag)
                    if v is not None and v >= 0:
                        candidates.append(v)
                        break
        if candidates:
            # UI often repeats labels in menus. Larger non-zero visible card value is usually the actual metric.
            out[key] = max(candidates)
    return out


def month_key(dt: datetime | None = None) -> str:
    dt = dt or datetime.now()
    return f"{dt.year}-{dt.month:02d}"


def upsert_monthly(data: dict[str, Any], platform: str, metrics: dict[str, Any], captured_at: str) -> None:
    month = month_key()
    row = next((r for r in data["monthly"] if r.get("month") == month and r.get("platform") == platform), None)
    if not row:
        row = {"id": f"collector-{platform}-{month}", "month": month, "platform": platform, "source": "daily-snapshot"}
        data["monthly"].append(row)
    for k, v in metrics.items():
        if k == "content_interactions":
            row["other_interactions"] = v
        else:
            row[k] = v
    row["captured_at"] = captured_at


def set_status(data: dict[str, Any], platform: str, ok: bool, message: str, captured_at: str) -> None:
    row = next((r for r in data["collector_status"] if r.get("platform") == platform), None)
    if not row:
        row = {"platform": platform}
        data["collector_status"].append(row)
    row.update({"ok": ok, "message": message, "last_run": captured_at})


def _login_required(text: str, url: str) -> bool:
    s = (text + " " + url).lower()
    hints = ["登入 facebook", "登入 instagram", "log into facebook", "log in to facebook", "log in to instagram", "登入 threads", "login"]
    return any(h in s for h in hints) and len(text) < 9000


def collect_platform(platform: str, cfg: dict[str, Any], headed: bool = False) -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    pcfg = cfg["platforms"][platform]
    url = pcfg.get("insights_url", "").strip()
    if not url:
        return {"ok": False, "message": "尚未設定 Insights 網址", "metrics": {}}

    stamp = datetime.now().astimezone().isoformat()
    day = datetime.now().strftime("%Y-%m-%d")
    target_dir = raw_dir() / day
    target_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            str(profile_dir()),
            headless=not headed,
            locale="zh-TW",
            accept_downloads=True,
            viewport={"width": 1500, "height": 1000},
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=90_000)
            page.wait_for_timeout(8000)
            # Let lazy-loaded insight cards appear.
            try:
                page.mouse.wheel(0, 1800)
                page.wait_for_timeout(1800)
                page.mouse.wheel(0, -1800)
                page.wait_for_timeout(1200)
            except Exception:
                pass
            text = page.locator("body").inner_text(timeout=20_000)
            final_url = page.url
            (target_dir / f"{platform}.txt").write_text(text, encoding="utf-8")
            (target_dir / f"{platform}.url.txt").write_text(final_url, encoding="utf-8")
            try:
                page.screenshot(path=str(target_dir / f"{platform}.png"), full_page=True)
            except Exception:
                pass
            if _login_required(text, final_url):
                return {"ok": False, "message": "登入狀態失效，請開啟小助手重新登入", "metrics": {}, "url": final_url, "captured_at": stamp}
            metrics = extract_metrics_from_text(text)
            if not metrics:
                return {"ok": False, "message": "頁面已開啟，但未辨識到指標；請用『設定頁面』重新定位 Insights 頁", "metrics": {}, "url": final_url, "captured_at": stamp}
            return {"ok": True, "message": f"擷取 {len(metrics)} 個指標", "metrics": metrics, "url": final_url, "captured_at": stamp}
        except Exception as e:
            return {"ok": False, "message": f"擷取失敗：{type(e).__name__}: {e}", "metrics": {}, "captured_at": stamp}
        finally:
            context.close()


def collect_all(headed: bool = False, do_git_sync: bool | None = None) -> dict[str, Any]:
    cfg = load_config()
    data = load_output_data(cfg)
    results = {}
    for platform in PLATFORMS:
        if not cfg.get("platforms", {}).get(platform, {}).get("enabled", True):
            continue
        result = collect_platform(platform, cfg, headed=headed)
        results[platform] = result
        stamp = result.get("captured_at", datetime.now().astimezone().isoformat())
        if result.get("ok"):
            upsert_monthly(data, platform, result.get("metrics", {}), stamp)
        set_status(data, platform, bool(result.get("ok")), result.get("message", ""), stamp)
    save_output_data(data, cfg)
    if do_git_sync is None:
        do_git_sync = bool(cfg.get("auto_git_sync"))
    git_result = git_sync(cfg) if do_git_sync else {"ok": True, "message": "未啟用 GitHub 自動同步"}
    return {"platforms": results, "git": git_result, "path": str(output_data_path(cfg))}


def setup_browser(platform: str) -> None:
    """Open a persistent browser for login/navigation and continuously expose its URL to the GUI."""
    from playwright.sync_api import sync_playwright
    cfg = load_config()
    url = cfg.get("platforms", {}).get(platform, {}).get("insights_url") or "https://business.facebook.com/"
    marker = runtime_dir() / f"current-url-{platform}.txt"
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            str(profile_dir()), headless=False, locale="zh-TW", viewport={"width": 1500, "height": 1000}
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=90_000)
        try:
            while True:
                pages = context.pages
                if not pages:
                    break
                page = pages[-1]
                marker.write_text(page.url, encoding="utf-8")
                time.sleep(1)
        except KeyboardInterrupt:
            pass
        finally:
            try: context.close()
            except Exception: pass


def current_setup_url(platform: str) -> str:
    p = runtime_dir() / f"current-url-{platform}.txt"
    return p.read_text(encoding="utf-8").strip() if p.exists() else ""


def norm_header(s: str) -> str:
    return normalize_text(s).replace("／", "").replace("/", "")


def detect_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except Exception:
        return ","


def read_tabular(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return []
        headers = [str(x or "").strip() for x in rows[0]]
        return [{headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))} for r in rows[1:] if any(x not in (None, "") for x in r)]

    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp950", "big5", "utf-16"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            pass
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    delim = detect_delimiter(text[:5000])
    return list(csv.DictReader(text.splitlines(), delimiter=delim))


def pick(row: dict[str, Any], key: str) -> Any:
    normalized = {norm_header(k): v for k, v in row.items()}
    for alias in HEADER_ALIASES.get(key, []):
        a = norm_header(alias)
        if a in normalized and normalized[a] not in (None, ""):
            return normalized[a]
    return ""


def to_number(v: Any) -> int | float:
    x = parse_human_number(str(v))
    return 0 if x is None else x


def infer_platform(path: Path, requested: str = "auto") -> str:
    if requested in PLATFORMS:
        return requested
    s = path.name.lower()
    if "instagram" in s or re.search(r"(^|[_\-])ig([_\-.]|$)", s): return "instagram"
    if "thread" in s: return "threads"
    return "facebook"


def parse_date(v: Any) -> str:
    if isinstance(v, datetime): return v.isoformat(timespec="minutes")
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try: return datetime.strptime(s[:19], fmt).isoformat(timespec="minutes")
        except Exception: pass
    return s


def activity_catalog(cfg: dict[str, Any]) -> list[dict[str, Any]]:
    repo = str(cfg.get("dashboard_repo_path", "")).strip()
    candidates = []
    if repo: candidates.append(Path(repo) / "data" / "manual-data.json")
    candidates.append(project_dir() / "data" / "manual-data.json")
    for p in candidates:
        if p.exists():
            try: return json.loads(p.read_text(encoding="utf-8")).get("activities", [])
            except Exception: pass
    return []


def classify_activity(text: str, date: str, activities: list[dict[str, Any]]) -> tuple[str, str]:
    nt = normalize_text(text)
    best_id, best_score = "", 0
    d = date[:10]
    for a in activities:
        score = 0
        terms = [a.get("name", ""), *(a.get("keywords") or [])]
        for term in terms:
            t = normalize_text(term)
            if t and t in nt: score += 5 if t == normalize_text(a.get("name", "")) else 2
        if d and a.get("start_date") and a.get("end_date") and a["start_date"] <= d <= a["end_date"]:
            score += 1
        if score > best_score:
            best_id, best_score = a.get("id", ""), score
    if best_score >= 5: return best_id, "high"
    if best_score >= 3: return best_id, "medium"
    if best_score >= 2: return best_id, "low"
    return "", "none"


def import_content_file(path: str, platform: str = "auto") -> dict[str, Any]:
    cfg = load_config()
    data = load_output_data(cfg)
    p = Path(path)
    rows = read_tabular(p)
    pf = infer_platform(p, platform)
    activities = activity_catalog(cfg)
    imported = 0
    existing = {x.get("fingerprint") for x in data["content"] if x.get("fingerprint")}
    for row in rows:
        text = str(pick(row, "text") or "").strip()
        date = parse_date(pick(row, "published_at"))
        url = str(pick(row, "url") or "").strip()
        if not text and not url: continue
        fp = hashlib.sha256(f"{pf}|{date}|{url}|{text[:500]}".encode("utf-8", errors="ignore")).hexdigest()[:20]
        if fp in existing: continue
        aid, conf = classify_activity(f"{text} {url}", date, activities)
        item = {
            "id": f"import-{fp}", "fingerprint": fp, "source": "meta-export", "platform": pf,
            "published_at": date, "content_type": str(pick(row, "content_type") or "貼文"),
            "url": url, "text": text, "views": to_number(pick(row, "views")), "reach": to_number(pick(row, "reach")),
            "likes": to_number(pick(row, "likes")), "comments": to_number(pick(row, "comments")),
            "shares": to_number(pick(row, "shares")), "saves": to_number(pick(row, "saves")),
            "clicks": to_number(pick(row, "clicks")), "other": 0,
            "activity_id": aid, "confidence": conf, "assignment_source": "auto", "included": conf not in ("low", "none")
        }
        data["content"].append(item); existing.add(fp); imported += 1

    # If the file contains post-level metrics, also fill missing monthly content counts.
    month_items = defaultdict(list)
    for item in data["content"]:
        if item.get("platform") != pf: continue
        m = str(item.get("published_at") or "")[:7]
        if len(m) == 7: month_items[m].append(item)
    for month, items in month_items.items():
        row = next((r for r in data["monthly"] if r.get("month") == month and r.get("platform") == pf), None)
        if not row:
            row = {"id": f"collector-{pf}-{month}", "month": month, "platform": pf, "source": "content-export"}
            data["monthly"].append(row)
        types = [normalize_text(x.get("content_type")) for x in items]
        row["posts"] = sum(1 for t in types if "reel" not in t and "限時" not in t and "story" not in t)
        row["reels"] = sum(1 for t in types if "reel" in t)
        row["stories"] = sum(1 for t in types if "限時" in t or "story" in t)
        # Do not overwrite account-level daily snapshot views/reach if they already exist.
        row.setdefault("views", sum(to_number(x.get("views")) for x in items))
        row.setdefault("reach", sum(to_number(x.get("reach")) for x in items))
        row["likes"] = sum(to_number(x.get("likes")) for x in items)
        row["comments"] = sum(to_number(x.get("comments")) for x in items)
        row["shares"] = sum(to_number(x.get("shares")) for x in items)
        row["saves"] = sum(to_number(x.get("saves")) for x in items)
        row["link_clicks"] = sum(to_number(x.get("clicks")) for x in items)
        row["content_export_at"] = datetime.now().astimezone().isoformat()
    save_output_data(data, cfg)
    return {"platform": pf, "rows": len(rows), "imported": imported, "path": str(output_data_path(cfg))}


def classify_topic(text: str) -> str:
    nt = normalize_text(text)
    best, score = "other", 0
    for key, (_, words) in TOPICS.items():
        if key == "other": continue
        s = sum(1 for w in words if normalize_text(w) in nt)
        if s > score: best, score = key, s
    return best


def iter_json_objects(obj: Any) -> Iterable[dict[str, Any]]:
    if isinstance(obj, dict):
        yield obj
        for v in obj.values(): yield from iter_json_objects(v)
    elif isinstance(obj, list):
        for v in obj: yield from iter_json_objects(v)


def extract_messages_from_json(obj: Any) -> list[dict[str, Any]]:
    msgs: list[dict[str, Any]] = []
    for d in iter_json_objects(obj):
        if isinstance(d.get("messages"), list):
            for m in d["messages"]:
                if not isinstance(m, dict): continue
                content = m.get("content") or m.get("text") or m.get("message")
                sender = m.get("sender_name") or m.get("sender") or m.get("from") or ""
                ts = m.get("timestamp_ms") or m.get("timestamp") or m.get("created_at") or ""
                if content:
                    msgs.append({"content": str(content), "sender": str(sender), "timestamp": ts})
    # de-dup recursive discoveries
    seen, out = set(), []
    for m in msgs:
        fp = hashlib.sha256(f"{m['sender']}|{m['timestamp']}|{m['content']}".encode("utf-8", errors="ignore")).hexdigest()
        if fp not in seen: seen.add(fp); out.append(m)
    return out


def load_json_files(path: Path) -> list[tuple[str, Any]]:
    out = []
    if path.is_file() and path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path) as z:
            for name in z.namelist():
                if name.lower().endswith(".json"):
                    try: out.append((name, json.loads(z.read(name).decode("utf-8"))))
                    except Exception: pass
    elif path.is_file() and path.suffix.lower() == ".json":
        try: out.append((path.name, json.loads(path.read_text(encoding="utf-8"))))
        except Exception: pass
    elif path.is_dir():
        for f in path.rglob("*.json"):
            try: out.append((str(f), json.loads(f.read_text(encoding="utf-8"))))
            except Exception: pass
    return out


def timestamp_to_date(ts: Any) -> str:
    try:
        if isinstance(ts, (int, float)):
            x = float(ts)
            if x > 10_000_000_000: x /= 1000
            return datetime.fromtimestamp(x).astimezone().strftime("%Y-%m-%d")
    except Exception: pass
    s = str(ts or "")
    return s[:10] if re.match(r"\d{4}-\d{2}-\d{2}", s) else datetime.now().strftime("%Y-%m-%d")


def import_private_messages(path: str, platform: str = "auto") -> dict[str, Any]:
    """Analyze locally. Only anonymous counts are written to collector-data.json; raw messages stay local."""
    cfg = load_config()
    data = load_output_data(cfg)
    p = Path(path)
    pf = infer_platform(p, platform)
    own_names = [normalize_text(x) for x in cfg.get("own_sender_names", []) if x]
    activities = activity_catalog(cfg)
    files = load_json_files(p)

    private_rows: list[dict[str, Any]] = []
    conversation_months: set[tuple[str, str]] = set()
    seen_messages: set[str] = set()
    messages_found = 0
    for source_name, obj in files:
        file_msgs = extract_messages_from_json(obj)
        messages_found += len(file_msgs)
        file_had_months: set[str] = set()
        for m in file_msgs:
            sender_n = normalize_text(m.get("sender"))
            if own_names and any(x and x in sender_n for x in own_names):
                continue
            date = timestamp_to_date(m.get("timestamp"))
            fp = hashlib.sha256(f"{pf}|{date}|{m.get('sender','')}|{m['content']}".encode("utf-8", errors="ignore")).hexdigest()[:20]
            if fp in seen_messages:
                continue
            seen_messages.add(fp)
            topic = classify_topic(m["content"])
            aid, conf = classify_activity(m["content"], date, activities)
            private_rows.append({
                "id": fp, "date": date, "platform": pf, "activity_id": aid, "confidence": conf, "topic": topic,
                "text": m["content"], "sender": m.get("sender", ""), "source_file": source_name, "included": True
            })
            file_had_months.add(date[:7])
        for month in file_had_months:
            conversation_months.add((source_name, month))

    private_path = private_dir() / f"messages-{pf}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    private_path.write_text(json.dumps(private_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    # Dashboard receives only anonymous aggregate counts.
    rebuild_private_aggregates(private_path)
    data = load_output_data(cfg)
    per_month_messages = Counter(r["date"][:7] for r in private_rows if r.get("included") is not False and len(r.get("date", "")) >= 7)
    per_month_conversations = Counter(month for _, month in conversation_months)
    for month in sorted(set(per_month_messages) | set(per_month_conversations)):
        row = next((r for r in data["monthly"] if r.get("month") == month and r.get("platform") == pf), None)
        if not row:
            row = {"id": f"collector-{pf}-{month}", "month": month, "platform": pf, "source": "message-export"}
            data["monthly"].append(row)
        row["inbox_messages"] = per_month_messages.get(month, 0)
        row["inbox_conversations"] = per_month_conversations.get(month, 0)
        row["message_export_at"] = datetime.now().astimezone().isoformat()
    save_output_data(data, cfg)
    return {
        "platform": pf, "json_files": len(files), "messages_found": messages_found,
        "incoming_analyzed": len(private_rows), "conversations_estimated": len(conversation_months),
        "private_detail": str(private_path), "dashboard_path": str(output_data_path(cfg))
    }

def latest_private_file() -> Path | None:
    files = sorted(private_dir().glob("messages-*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
    return files[0] if files else None


def load_private_rows(path: str | Path | None = None) -> tuple[Path | None, list[dict[str, Any]]]:
    p = Path(path) if path else latest_private_file()
    if not p or not p.exists(): return None, []
    try: return p, json.loads(p.read_text(encoding="utf-8"))
    except Exception: return p, []


def rebuild_private_aggregates(path: str | Path | None = None) -> dict[str, Any]:
    p, rows = load_private_rows(path)
    if not p: return {"ok": False, "message": "找不到私訊本機分析檔"}
    cfg = load_config(); data = load_output_data(cfg)
    bucket: dict[tuple[str,str,str,str], int] = defaultdict(int)
    touched: set[tuple[str,str]] = set()
    for r in rows:
        if r.get("included") is False: continue
        pf = r.get("platform") or "other"
        month = str(r.get("date") or "")[:7]
        if len(month) != 7: continue
        aid = r.get("activity_id") or ""
        topic = r.get("topic") or "other"
        touched.add((pf, month)); bucket[(pf,month,aid,topic)] += 1
    data["inquiry_aggregates"] = [r for r in data["inquiry_aggregates"] if (r.get("platform"),r.get("month")) not in touched]
    for (pf,month,aid,topic),count in sorted(bucket.items()):
        data["inquiry_aggregates"].append({"id":f"inq-{pf}-{month}-{aid or 'unassigned'}-{topic}","month":month,"platform":pf,"activity_id":aid,"topic":topic,"count":count,"source":"private-local-analysis"})
    save_output_data(data,cfg)
    return {"ok": True, "message": f"已依本機校正結果重建 {sum(bucket.values())} 則匿名統計", "path": str(output_data_path(cfg))}


def save_private_rows(rows: list[dict[str, Any]], path: str | Path | None = None) -> dict[str, Any]:
    p = Path(path) if path else latest_private_file()
    if not p: return {"ok": False, "message": "找不到私訊本機分析檔"}
    p.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return rebuild_private_aggregates(p)

def git_sync(cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    cfg = cfg or load_config()
    repo = str(cfg.get("dashboard_repo_path", "")).strip()
    if not repo:
        return {"ok": False, "message": "尚未設定本機 GitHub Repository 資料夾"}
    repo_path = Path(repo)
    if not (repo_path / ".git").exists():
        return {"ok": False, "message": "指定資料夾不是 Git Repository；請先用 GitHub Desktop Clone"}
    try:
        subprocess.run(["git", "-C", str(repo_path), "pull", "--rebase", "--autostash"], check=True, capture_output=True, text=True, timeout=120)
        subprocess.run(["git", "-C", str(repo_path), "add", "data/collector-data.json"], check=True, capture_output=True, text=True, timeout=60)
        diff = subprocess.run(["git", "-C", str(repo_path), "diff", "--cached", "--quiet"], capture_output=True)
        if diff.returncode == 0:
            return {"ok": True, "message": "資料無變更，不需推送"}
        subprocess.run(["git", "-C", str(repo_path), "commit", "-m", f"Update social insights {datetime.now().strftime('%Y-%m-%d %H:%M')}"], check=True, capture_output=True, text=True, timeout=60)
        subprocess.run(["git", "-C", str(repo_path), "push"], check=True, capture_output=True, text=True, timeout=120)
        return {"ok": True, "message": "已更新 GitHub，Pages 將自動重新部署"}
    except FileNotFoundError:
        return {"ok": False, "message": "找不到 Git。建議安裝 GitHub Desktop 後再啟用自動同步"}
    except subprocess.CalledProcessError as e:
        msg = (e.stderr or e.stdout or str(e)).strip()
        return {"ok": False, "message": "Git 同步失敗：" + msg[-500:]}
    except Exception as e:
        return {"ok": False, "message": f"Git 同步失敗：{e}"}


def install_task(time_hhmm: str) -> dict[str, Any]:
    if os.name != "nt":
        return {"ok": False, "message": "每日工作排程安裝只支援 Windows"}
    try:
        hh, mm = [int(x) for x in time_hhmm.split(":")]
        st = f"{hh:02d}:{mm:02d}"
    except Exception:
        return {"ok": False, "message": "時間格式需為 HH:MM"}
    runner = Path(__file__).resolve().parent / "run_scheduled.bat"
    cmd = ["schtasks", "/Create", "/F", "/SC", "DAILY", "/ST", st, "/TN", "Social Impact Collector", "/TR", f'"{runner}"']
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if r.returncode == 0:
            return {"ok": True, "message": f"已建立每日 {st} 自動蒐集工作"}
        return {"ok": False, "message": (r.stderr or r.stdout).strip()}
    except Exception as e:
        return {"ok": False, "message": str(e)}


def remove_task() -> dict[str, Any]:
    if os.name != "nt": return {"ok": False, "message": "只支援 Windows"}
    r = subprocess.run(["schtasks", "/Delete", "/F", "/TN", "Social Impact Collector"], capture_output=True, text=True)
    return {"ok": r.returncode == 0, "message": (r.stdout or r.stderr).strip()}
